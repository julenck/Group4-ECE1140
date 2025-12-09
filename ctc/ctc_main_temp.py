
import json, time, os
from datetime import datetime
from .ctc_main_helper_functions import JSONFileWatcher
from watchdog.observers import Observer
from .track.map import route_lookup_via_station, route_lookup_via_id, route_info

def safe_json_read(file_path, max_retries=3, delay=0.1):
    """Safely read JSON with retry logic to handle race conditions."""
    for attempt in range(max_retries):
        try:
            with open(file_path, "r") as f:
                return json.load(f)
        except json.JSONDecodeError as e:
            if attempt < max_retries - 1:
                time.sleep(delay * (2 ** attempt))  # exponential backoff
            else:
                print(f"[CTC] Failed to read {file_path} after {max_retries} attempts: {e}")
                return None
        except Exception as e:
            if attempt < max_retries - 1:
                time.sleep(delay * (2 ** attempt))
            else:
                print(f"[CTC] Error reading {file_path}: {e}")
                return None
    return None

def safe_json_write(file_path, data, max_retries=3, delay=0.1):
    """Safely write JSON with retry logic to handle race conditions."""
    for attempt in range(max_retries):
        try:
            with open(file_path, "w") as f:
                json.dump(data, f, indent=4)
            return True
        except Exception as e:
            if attempt < max_retries - 1:
                time.sleep(delay * (2 ** attempt))
            else:
                print(f"[CTC] Failed to write {file_path} after {max_retries} attempts: {e}")
                return False
    return False

def track_update_handler(new_data, train, data_file_ctc_data):
    try:
        train_pos = new_data["Trains"][train]["Train Position"]
        train_state = new_data["Trains"][train]["Train State"]
        print(f"train position {train_pos}")
        data = safe_json_read(data_file_ctc_data)
        if data:
            data["Dispatcher"]["Trains"][train]["Position"] = train_pos
            data["Dispatcher"]["Trains"][train]["State"] = train_state
            safe_json_write(data_file_ctc_data, data)
    except Exception as e:
        print(f"Train position data missing: {e}")

def dispatch_train(train, line, station, arrival_time_str,
                   data_file_ctc_data='ctc_data.json',
                   data_file_track_cont='../ctc_track_controller.json',
                   dwell_time_s=10,
                   is_single_station_dispatch=False):
    print(f"\n[DISPATCH_TRAIN] Called with: train={train}, line={line}, station={station}, is_single_station_dispatch={is_single_station_dispatch}")
    # Resolve file paths relative to the ctc package base directory when not absolute
    BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    if not os.path.isabs(data_file_ctc_data):
        # ensure ctc_data lives in the project base dir
        data_file_ctc_data = os.path.join(BASE_DIR, os.path.basename(data_file_ctc_data))
    if not os.path.isabs(data_file_track_cont):
        # avoid honoring leading '..' in the default path; place track controller file in project base dir
        data_file_track_cont = os.path.join(BASE_DIR, os.path.basename(data_file_track_cont))

    # normalize to absolute paths and report locations for debugging
    data_file_ctc_data = os.path.abspath(data_file_ctc_data)
    data_file_track_cont = os.path.abspath(data_file_track_cont)
    print(f"CTC data file -> {data_file_ctc_data}")
    print(f"Track controller file -> {data_file_track_cont}")

    # Only reset JSON files if NOT in single-station dispatch mode
    # In single-station mode (used by dispatch_schedule), we preserve state across calls
    if not is_single_station_dispatch:
        print("Resetting CTC JSON files to default state...")
        
        # Reset ctc_data.json
        default_ctc_data = {
            "Dispatcher": {
                "Trains": {}
            }
        }
        for i in range(1, 6):
            tname = f"Train {i}"
            default_ctc_data["Dispatcher"]["Trains"][tname] = {
                "Line": "",
                "Suggested Speed": "",
                "Authority": "",
                "Station Destination": "",
                "Arrival Time": "",
                "Position": "",
                "State": "",
                "Current Station": ""
            }
        safe_json_write(data_file_ctc_data, default_ctc_data)
        print("ctc_data.json reset complete")
        
        # Reset ctc_track_controller.json
        default_track = {"Trains": {}}
        for i in range(1, 6):
            tname = f"Train {i}"
            default_track["Trains"][tname] = {
                "Active": 0,
                "Suggested Speed": 0,
                "Suggested Authority": 0,
                "Train Position": 0,
                "Train State": 0
            }
        safe_json_write(data_file_track_cont, default_track)
        print("ctc_track_controller.json reset complete")
    else:
        print("[SINGLE_STATION_DISPATCH] Skipping file reset to preserve train state across multiple stations")

    # Ensure the track controller file exists (create minimal structure if missing)
    if not os.path.exists(data_file_track_cont):
        try:
            with open(data_file_track_cont, 'w') as f:
                json.dump({"Trains": {}}, f, indent=4)
        except Exception:
            # If we cannot create the file, let the watcher raise a clear error
            pass

    # Ensure the ctc data file exists with minimal structure so UI updates succeed
    if not os.path.exists(data_file_ctc_data):
        try:
            with open(data_file_ctc_data, 'w') as f:
                json.dump({"Dispatcher": {"Trains": {}}}, f, indent=4)
        except Exception:
            pass

    # Ensure both files contain default train entries expected by the UI/dispatcher
    def _ensure_train_entries():
        trains = [f"Train {i}" for i in range(1, 6)]
        # ctc_data: ensure Dispatcher->Trains has entries
        ctc_data = safe_json_read(data_file_ctc_data)
        if not ctc_data:
            ctc_data = {"Dispatcher": {"Trains": {}}}

        dispatcher_trains = ctc_data.setdefault("Dispatcher", {}).setdefault("Trains", {})
        for t in trains:
            dispatcher_trains.setdefault(t, {
                "Line": "",
                "Suggested Speed": "",
                "Authority": "",
                "Station Destination": "",
                "Arrival Time": "",
                "Position": "",
                "State": "",
                "Current Station": ""
            })

        safe_json_write(data_file_ctc_data, ctc_data)

        # track controller file: ensure Trains has entries
        track_updates = safe_json_read(data_file_track_cont)
        if not track_updates:
            track_updates = {"Trains": {}}

        trains_dict = track_updates.setdefault("Trains", {})
        for t in trains:
            trains_dict.setdefault(t, {
                "Active": 0,
                "Suggested Authority": 0,
                "Suggested Speed": 0,
                "Train Position": None,
                "Train State": ""
            })

        safe_json_write(data_file_track_cont, track_updates)

    _ensure_train_entries()

    dest_id = route_lookup_via_station[station]["id"]
    print(f"dest id ={dest_id}")
    
    # Build route sequence lookup by index (0, 1, 2, ..., 20)
    route_by_sequence = {}
    for i in range(len(route_info["ID"])):
        route_by_sequence[i] = {
            "name": route_info["Name"][i],
            "meters_to_next": route_info["Meters to next"][i],
            "block": route_info["Block"][i]
        }
    
    total_dist = 0
    for i in range(dest_id+1):
        total_dist += route_by_sequence[i]["meters_to_next"]
    print(f"dist to destination = {total_dist}")
    now = datetime.now()
    print(now)
    # In single-station dispatch mode, only account for dwell at the final destination
    # In regular mode, account for dwell at each intermediate station
    if is_single_station_dispatch:
        total_dwell_time_s = dwell_time_s  # Only dwell at final destination
    else:
        total_dwell_time_s = dwell_time_s * dest_id  # Dwell at each intermediate station
    print(f"total_dwell_time_s = {total_dwell_time_s} (is_single_station_dispatch={is_single_station_dispatch})")
    arrival_time = datetime.strptime(arrival_time_str, "%H:%M")
    arrival_time = arrival_time.replace(year=now.year, month=now.month, day=now.day)
    print(arrival_time)
    total_time = arrival_time - now
    total_time_s = total_time.total_seconds()
    print(total_time_s)
    time_to_drive_s = total_time_s - total_dwell_time_s
    print(time_to_drive_s)
    speed_meters_s = int(total_dist / time_to_drive_s)
    speed_mph = int(speed_meters_s * 2.237)
    print(f"speed in m/s = {speed_meters_s}")
    print(f"speed in mph = {speed_mph}")
    print(f"[DEBUG] Calculated values: speed_mph={speed_mph}, will write to ctc_data as 'Suggested Speed'")
    train_pos = None
    train_state = None
    current_station_id = None  # Track which station the train is currently at

    def _track_update_handler(new_data):
        nonlocal train_pos, train_state
        try:
            train_pos = new_data["Trains"][train]["Train Position"]
            train_state = new_data["Trains"][train]["Train State"]
            print(f"train position {train_pos}")
            data = safe_json_read(data_file_ctc_data)
            if data:
                data["Dispatcher"]["Trains"][train]["Position"] = train_pos
                data["Dispatcher"]["Trains"][train]["State"] = train_state
                safe_json_write(data_file_ctc_data, data)
        except Exception as e:
            print(f"Train position data missing: {e}")

    watcher = JSONFileWatcher(data_file_track_cont, _track_update_handler)
    observer = Observer()
    observer.schedule(watcher, os.path.dirname(data_file_track_cont), recursive=False)
    observer.start()

    try:
        # In single-station dispatch mode, only dispatch to the destination station directly
        # Do not loop through intermediate stations
        if is_single_station_dispatch:
            loop_range = [dest_id]  # Only dispatch to the destination
        else:
            loop_range = range(dest_id + 1)  # Full journey from start to destination
        
        for i in loop_range:
            test = route_by_sequence[i]["name"]
            print(f"[MAIN_LOOP] Station {i}: {test}, is_single_station_dispatch={is_single_station_dispatch}, destination={station}")
            print(f"next station ={test}")
            # Authority for trains is based on the destination station's index
            # The authority is the distance FROM the destination station TO the next station
            authority_meters = route_info["Meters to next"][dest_id]
            authority_yards = int(authority_meters * 1.094)
            print(f"authority in meters = {authority_meters} (from track map at destination station index {dest_id})")
            next_station_loc = route_by_sequence[i]["block"]
            print(f"next station loc = {next_station_loc}")
            
            # Update CTC data with dispatch information
            data = safe_json_read(data_file_ctc_data)
            if data:
                data["Dispatcher"]["Trains"][train]["Line"] = line
                data["Dispatcher"]["Trains"][train]["Authority"] = authority_yards
                data["Dispatcher"]["Trains"][train]["Suggested Speed"] = speed_mph
                data["Dispatcher"]["Trains"][train]["Station Destination"] = station
                data["Dispatcher"]["Trains"][train]["Arrival Time"] = arrival_time_str
                print(f"[DEBUG] Writing to ctc_data: Authority={authority_yards} yards, Speed={speed_mph} mph")
                safe_json_write(data_file_ctc_data, data)
            
            # Update track controller
            updates = safe_json_read(data_file_track_cont)
            if updates:
                updates["Trains"][train]["Active"] = 1
                updates["Trains"][train]["Suggested Authority"] = authority_meters
                updates["Trains"][train]["Suggested Speed"] = speed_meters_s
                safe_json_write(data_file_track_cont, updates)
            
            # Wait for train to reach station
            print(f"[WAIT] Waiting for train to reach block {next_station_loc}, current train_pos={train_pos}")
            while train_pos is None or train_pos != next_station_loc:
                time.sleep(0.5)
                print(f"[WAIT] Still waiting... train_pos={train_pos}, target={next_station_loc}")
            print(f"[REACHED] Train has reached block {next_station_loc}")
            
            current_station = test
            # Update current station
            data = safe_json_read(data_file_ctc_data)
            if data:
                data["Dispatcher"]["Trains"][train]["Current Station"] = current_station
                data["Dispatcher"]["Trains"][train]["Authority"] = authority_yards
                data["Dispatcher"]["Trains"][train]["Suggested Speed"] = speed_mph
                safe_json_write(data_file_ctc_data, data)
            
            print(f"Train arrived at {test}")
            
            # Wait for track controller to set Active = 0 (when authority exhausted)
            print(f"[CTC] Waiting for {train} to stop at station (authority exhausted)...")
            while True:
                updates = safe_json_read(data_file_track_cont)
                if updates and updates["Trains"][train]["Active"] == 0:
                    print(f"[CTC] {train} has stopped at station (Active=0 by track controller)")
                    break
                time.sleep(0.5)
            
            # Wait for dwell time
            time.sleep(dwell_time_s)
            time.sleep(0.5)
            
            # Keep current station visible while dwelling (Active = 0)
            # When we reactivate the train (set Active = 1), clear the current station
            
            # If not at final destination, set Active = 1 with new authority for next leg
            if test != station:
                # Authority is always based on the final destination station
                next_authority_meters = route_info["Meters to next"][dest_id]
                next_authority_yards = int(next_authority_meters * 1.094)
                
                # Update ctc_data.json with new authority and clear current station
                data = safe_json_read(data_file_ctc_data)
                if data:
                    data["Dispatcher"]["Trains"][train]["Authority"] = next_authority_yards
                    data["Dispatcher"]["Trains"][train]["Current Station"] = ""  # Clear when moving
                    safe_json_write(data_file_ctc_data, data)
                
                # Update track controller with new authority
                updates = safe_json_read(data_file_track_cont)
                if updates:
                    updates["Trains"][train]["Active"] = 1
                    updates["Trains"][train]["Suggested Authority"] = next_authority_meters
                    updates["Trains"][train]["Suggested Speed"] = speed_meters_s
                    safe_json_write(data_file_track_cont, updates)
                
                print(f"[CTC] Reactivated {train} with authority={next_authority_meters}m for next leg")
            
            if test == station:
                print("train at destination")
                break
    except KeyboardInterrupt:
        print("stopping observer")
    finally:
        observer.stop()
        observer.join()

def dispatch_schedule(schedule_file_path,
                      data_file_ctc_data='ctc_data.json',
                      data_file_track_cont='../ctc_track_controller.json',
                      dwell_time_s=10):
    """
    Dispatch trains automatically from a schedule file.
    Schedule format: CSV with columns: Line, Station, (Arrival Time or minutes offset)
    Can also include: Train (optional), Block Number (optional)
    
    IMPORTANT: Only the provided schedule_file_path is used. No fallback or test schedules are attempted.
    """
    # Validate that schedule_file_path is provided and not None
    if not schedule_file_path:
        print("[SCHEDULE] ERROR: No schedule file path provided. Must use uploaded schedule only.")
        raise ValueError("schedule_file_path cannot be None or empty")
    
    # Validate that the schedule file exists
    if not os.path.exists(schedule_file_path):
        print(f"[SCHEDULE] ERROR: Schedule file does not exist: {schedule_file_path}")
        raise FileNotFoundError(f"Schedule file not found: {schedule_file_path}")
    
    print(f"[SCHEDULE] Using schedule file: {schedule_file_path}")
    
    # Resolve file paths
    BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    if not os.path.isabs(data_file_ctc_data):
        data_file_ctc_data = os.path.join(BASE_DIR, os.path.basename(data_file_ctc_data))
    if not os.path.isabs(data_file_track_cont):
        data_file_track_cont = os.path.join(BASE_DIR, os.path.basename(data_file_track_cont))

    data_file_ctc_data = os.path.abspath(data_file_ctc_data)
    data_file_track_cont = os.path.abspath(data_file_track_cont)
    
    print(f"[SCHEDULE] CTC data file -> {data_file_ctc_data}")
    print(f"[SCHEDULE] Track controller file -> {data_file_track_cont}")

    # Load schedule from file
    try:
        # Support CSV, JSON, and Excel formats
        if schedule_file_path.endswith('.xlsx') or schedule_file_path.endswith('.xls'):
            try:
                import openpyxl
                print("[SCHEDULE] Loading Excel file with openpyxl")
                wb = openpyxl.load_workbook(schedule_file_path)
                ws = wb.active
                
                # Get headers from first row
                headers = []
                for cell in ws[1]:
                    headers.append(cell.value)
                
                # Get data rows
                schedule = []
                for row in ws.iter_rows(min_row=2, values_only=False):
                    row_data = {}
                    for i, header in enumerate(headers):
                        if i < len(row):
                            row_data[header] = row[i].value
                    schedule.append(row_data)
                print(f"[SCHEDULE] Loaded {len(schedule)} trains from Excel file")
            except ImportError:
                print("[SCHEDULE] openpyxl not installed, trying pandas")
                try:
                    import pandas as pd
                    df = pd.read_excel(schedule_file_path)
                    schedule = df.to_dict('records')
                    print(f"[SCHEDULE] Loaded {len(schedule)} trains from Excel file with pandas")
                except ImportError:
                    print("[SCHEDULE] pandas not installed either, cannot read Excel file")
                    raise ImportError("Need openpyxl or pandas to read Excel files")
        elif schedule_file_path.endswith('.csv'):
            import csv
            with open(schedule_file_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                schedule = list(reader)
            print(f"[SCHEDULE] Loaded {len(schedule)} trains from CSV file")
        elif schedule_file_path.endswith('.json'):
            with open(schedule_file_path, 'r') as f:
                schedule = json.load(f)
            print(f"[SCHEDULE] Loaded {len(schedule)} trains from JSON file")
        else:
            # Try to detect format
            try:
                with open(schedule_file_path, 'r') as f:
                    schedule = json.load(f)
                print(f"[SCHEDULE] Loaded {len(schedule)} trains from JSON file (detected)")
            except:
                with open(schedule_file_path, 'r', encoding='utf-8') as f:
                    import csv
                    reader = csv.DictReader(f)
                    schedule = list(reader)
                print(f"[SCHEDULE] Loaded {len(schedule)} trains from CSV file (detected)")
    except Exception as e:
        print(f"[SCHEDULE] Error loading schedule file: {e}")
        import traceback
        traceback.print_exc()
        return

    print(f"[SCHEDULE] Loaded {len(schedule)} entries from schedule")
    
    # Helper function to get field with flexible column names
    def get_field(entry, *possible_names):
        """Get a field from entry dict, trying multiple possible column names (case-insensitive)"""
        if entry is None:
            return None
        for key in entry.keys():
            if key is not None:
                # Normalize the key for comparison
                normalized_key = str(key).strip().lower()
                for name in possible_names:
                    if normalized_key == name.lower():
                        value = entry[key]
                        # Strip whitespace from string values
                        if isinstance(value, str):
                            return value.strip()
                        return value
        return None
    
    # Dispatch each train in the schedule
    # Group by train if possible, or create individual dispatches
    current_train = None
    current_line = None
    current_stations = []
    
    for i, entry in enumerate(schedule):
        try:
            # Extract fields
            train = get_field(entry, "Train", "Train ID", "TrainID", "Train Number", "TrainNumber")
            line = get_field(entry, "Line", "Line Name", "LineName")
            station = get_field(entry, "Station", "Station Name", "Destination", "StationName")
            
            # Try to get arrival time - could be in HH:MM format or minutes offset
            arrival_time = get_field(entry, "Arrival Time", "ArrivalTime", "Arrival", "Time")
            minutes_offset = get_field(entry, "Train 1 Arrival Time at Station (in minutes from current time)",
                                      "Minutes", "Offset", "Minutes Offset")
            
            # Parse minutes offset if present
            if minutes_offset is not None and not arrival_time:
                try:
                    minutes_int = int(minutes_offset) if isinstance(minutes_offset, (int, float)) else int(str(minutes_offset))
                    # Convert minutes offset to HH:MM format
                    from datetime import datetime, timedelta
                    now = datetime.now()
                    arrival_dt = now + timedelta(minutes=minutes_int)
                    arrival_time = arrival_dt.strftime("%H:%M")
                    print(f"[SCHEDULE] Converted {minutes_offset} minutes to {arrival_time}")
                except Exception as e:
                    print(f"[SCHEDULE] Error converting minutes offset: {e}")
            
            # Use Train 1 if no train specified
            if not train:
                train = "Train 1"
            
            print(f"[SCHEDULE] Entry {i+1}: Train={train}, Line={line}, Station={station}, Time={arrival_time}")
            
            # Check all required fields
            if not all([line, station, arrival_time]):
                print(f"[SCHEDULE] Skipping entry {i+1}: Missing required fields (line={line}, station={station}, time={arrival_time})")
                continue
            
            print(f"\n[SCHEDULE] Dispatching Train {i+1}/{len(schedule)}: {train} to {station} at {arrival_time}")
            dispatch_train(
                train=train,
                line=line,
                station=station,
                arrival_time_str=arrival_time,
                data_file_ctc_data=data_file_ctc_data,
                data_file_track_cont=data_file_track_cont,
                dwell_time_s=dwell_time_s,
                is_single_station_dispatch=True
            )
            print(f"[SCHEDULE] Train {train} dispatch complete")
        except Exception as e:
            print(f"[SCHEDULE] Error dispatching entry {i+1}: {e}")
            import traceback
            traceback.print_exc()
            continue


if __name__ == "__main__":
    # Example usage: read from ctc_ui_inputs.json and dispatch
    data_file_ui_inputs = 'ctc_ui_inputs.json'
    with open(data_file_ui_inputs, "r") as f_inputs:
        inputs = json.load(f_inputs)
    dispatch_train(
        train=inputs.get("Train"),
        line=inputs.get("Line"),
        station=inputs.get("Station"),
        arrival_time_str=inputs.get("Arrival Time")
    )

